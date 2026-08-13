#!/usr/bin/env python3
"""台帳xlsx → data/products.json (ja/en/zh 表示名・スペック行・フィルタタグ付き)
使い方: python3 scripts/build_data.py "<台帳xlsxパス>"
写真シート(写真/パール写真/DIA写真(C)/DIA写真(S))のアンカー位置から品番→画像と
出所シート(=コレクション判定)を得る。管理台帳シートが商品マスタ。
"""
import sys, os, re, json, zipfile, collections, csv
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    "~/Downloads/05_ジュエリーLC・GReEN/商品台帳_最新 (1).xlsx")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 最新カタログCSV(存在すれば品番の絞り込み+ct上書きに使用)
CATALOG_CSV = os.path.join(ROOT, "data", "catalog_20260813.csv")

def load_catalog():
    if not os.path.exists(CATALOG_CSV):
        return None
    m = {}
    for r in csv.DictReader(open(CATALOG_CSV)):
        m.setdefault(r["code"].strip(), []).append(r)
    return m

PHOTO_SHEETS = ["写真", "パール写真", "DIA写真(C)", "DIA写真(S)"]
SHEET_COLLECTION = {"写真": "jb", "パール写真": "pearl", "DIA写真(C)": "dia", "DIA写真(S)": "dia"}

# ---------- 写真シート: 品番→出所シート ----------
def map_photo_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    z = zipfile.ZipFile(path)
    sheet_drawing = {}
    for n in z.namelist():
        m = re.match(r"xl/worksheets/_rels/sheet(\d+)\.xml\.rels$", n)
        if m:
            d = z.read(n).decode()
            dm = re.search(r'Target="\.\./drawings/(drawing\d+)\.xml"', d)
            if dm:
                sheet_drawing[wb.sheetnames[int(m.group(1)) - 1]] = dm.group(1)
    code_sheet = {}
    for sname in PHOTO_SHEETS:
        if sname not in wb.sheetnames or sname not in sheet_drawing:
            continue
        ws = wb[sname]
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip():
                    code_sheet.setdefault(str(c.value).strip(), sname)
    return code_sheet

# ---------- 商品名パース ----------
TYPO = {"PNEDANT": "PENDANT", "PANDANT": "PENDANT", "NECLLACE": "NECKLACE",
        "NECKLES": "NECKLACE", "NEckLACE": "NECKLACE",
        "DIAMONDS": "DIA", "DIAMOND": "DIA"}
WORDS = {"DIAMONDS", "DIAMOND", "DIA", "ROSE", "CUT", "TREAT", "TREATED", "BLUE",
         "YELLOW", "PINK", "EMERALD", "AKOYA", "BLACK", "WHITE", "SOUTH", "SEA",
         "PEARL", "FRESHWATER", "PURPLE", "METAL", "CROSS", "FLOWER", "HORSE",
         "SHOE", "PENDANT", "NECKLACE", "PIERCE", "RING", "EAR", "CUFF",
         "BRACELET", "BAROQUE", "NATURAL", "MULTI"}
SIZE_RE = re.compile(r"^\d+(\.\d+)?([-x×]\d+(\.\d+)?)?(mm|cm|g)?$", re.I)

GEMS = [  # (検出キー列, ja, en, zh) 上から順に最長一致
    (["ROSE", "CUT"],            "ローズカットダイヤモンド", "Rose-cut Diamond", "玫瑰式切割钻石"),
    (["TREAT", "BLUE"],          "トリートブルーダイヤモンド", "Treated Blue Diamond", "处理蓝钻"),
    (["YELLOW", "DIA"],          "イエローダイヤモンド", "Yellow Diamond", "黄钻"),
    (["PINK", "DIA"],            "ピンクダイヤモンド", "Pink Diamond", "粉钻"),
    (["EMERALD"],                "エメラルド＆ダイヤモンド", "Emerald & Diamond", "祖母绿与钻石"),
    (["WHITE", "SOUTH"],         "南洋白蝶パール", "White South Sea Pearl", "南洋白珍珠"),
    (["BLACK", "PEARL"],         "ブラックパール", "Black Pearl", "黑珍珠"),
    (["AKOYA"],                  "アコヤパール", "Akoya Pearl", "Akoya珍珠"),
    (["FRESHWATER"],             "淡水パール", "Freshwater Pearl", "淡水珍珠"),
    (["PURPLE"],                 "パープルパール", "Purple Pearl", "紫珍珠"),
    (["DIA"],                    "ダイヤモンド", "Diamond", "钻石"),
    (["METAL"],                  "メタル", "Metal", "金属"),
]
MOTIFS = {"CROSS": ("クロス", "Cross", "十字"), "FLOWER": ("フラワー", "Flower", "花朵"),
          "HORSE": ("ホースシュー", "Horseshoe", "马蹄"), "BAROQUE": ("バロック", "Baroque", "巴洛克"),
          "NATURAL": ("ナチュラル", "Natural", "天然"), "MULTI": ("マルチカラー", "Multicolor", "多彩")}
TYPES = {"PD NECKLES": ("ペンダントネックレス", "Pendant Necklace", "吊坠项链"),
         "PENDANT":    ("ペンダントトップ", "Pendant Top", "吊坠"),
         "NECKLES":    ("ネックレス", "Necklace", "项链"),
         "PIERCE":     ("ピアス", "Pierced Earrings", "耳环"),
         "EAR CUFF":   ("イヤーカフ", "Ear Cuff", "耳骨夹"),
         "RING":       ("リング", "Ring", "戒指"),
         "BRACELET":   ("ブレスレット", "Bracelet", "手链")}
CAT_GROUP = {"PD NECKLES": "pendant", "PENDANT": "pendant", "NECKLES": "necklace",
             "PIERCE": "earrings", "EAR CUFF": "earrings", "RING": "ring", "BRACELET": "bracelet"}

def parse_name(raw, cat):
    toks = [TYPO.get(t.upper(), t.upper()) for t in raw.split()]
    sizes, words = [], []
    for t in toks:
        if SIZE_RE.match(t):
            sizes.append(t.lower())
        elif t in WORDS:
            words.append(t)
        # それ以外(数字/ハイフン混じりの社内コード)は捨てる
    gem = None
    for keys, ja, en, zh in GEMS:
        if all(k in words for k in keys):
            gem = (ja, en, zh); break
    motif = None
    for k, v in MOTIFS.items():
        if k in words and (gem is None or k not in ("BAROQUE", "NATURAL", "MULTI") or True):
            motif = v; break
    ty = TYPES.get(cat, ("ジュエリー", "Jewellery", "珠宝"))
    # 名前タイプ: 名称にNECKLACE系が無くカテゴリPENDANTならペンダントトップ表記のまま
    def nm(i):
        parts = []
        if gem: parts.append(gem[i])
        if motif: parts.append(motif[i])
        parts.append(ty[i])
        return "".join(parts) if i == 2 else " ".join(parts)
    # サイズ表記整形 (単位なし数値レンジはmm扱い=パール)
    fs = []
    for s in sizes:
        if re.match(r"^\d+(\.\d+)?([-x×]\d+(\.\d+)?)?$", s):
            s += "mm"
        fs.append(s)
    return {"ja": nm(0), "en": nm(1), "zh": nm(2)}, gem, ty, fs

def spec_line(gem, ty, mat, ct, sizes):
    def one(i, sep):
        p = [ty[i], mat] if mat else [ty[i]]
        g = gem[i] if gem else None
        extra = ""
        if ct:
            extra = f" {ct}ct"
        elif sizes:
            mm = [s for s in sizes if s.endswith("mm")]
            if mm: extra = f" {mm[0]}"
        if g: p.append(g + extra)
        ln = [s for s in sizes if s.endswith("cm")]
        if ln: p.append(("長さ" if i == 0 else ("length " if i == 1 else "长度 ")) + ln[0])
        return sep.join(p)
    return {"ja": one(0, "、"), "en": one(1, ", "), "zh": one(2, "、")}

def badge_of(sheet, ja, gemgrp, price):
    """マーチャンダイジングバッジ(仮ルール・要調整): 新作>人気>定番"""
    if sheet == "写真":
        return "new"
    if any(k in ja for k in ("クロス", "フラワー", "ホースシュー", "イエロー", "ピンク", "トリートブルー", "エメラルド", "ローズカット")):
        return "popular"
    if gemgrp == "dia" and price and 250000 <= price <= 400000:
        return "popular"
    if "アコヤ" in ja:
        return "classic"
    return None

def main():
    code_sheet = map_photo_sheets(XLSX)
    catalog = load_catalog()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["管理台帳"]
    products = collections.OrderedDict()
    for row in ws.iter_rows(min_row=3):
        v = [c.value for c in row]
        if not v[2]:
            continue
        code = str(v[2]).strip()
        if catalog is not None and code not in catalog:
            continue  # 最新カタログに無い品番は掲載しない
        price = v[7] if isinstance(v[7], (int, float)) else None
        if code in products:
            p = products[code]
            p["stock"] += 1
            if price and price != p["price"]:
                p["price_max"] = max(p.get("price_max", p["price"]), price)
                p["price"] = min(p["price"], price)
            continue
        raw = str(v[3]).strip() if v[3] else ""
        cat = str(v[4]).strip() if v[4] else ""
        mat = str(v[5]).strip() if v[5] else ""
        ct = v[6]
        if isinstance(ct, str):
            try: ct = float(ct)
            except ValueError: ct = None
        if isinstance(ct, float) and ct == 0: ct = None
        # カタログCSVのct表記(複合表記含む)があれば優先
        if catalog is not None:
            c_ct = (catalog[code][0].get("ct") or "").strip()
            if c_ct:
                c_ct = c_ct[:-2] if c_ct.endswith("ct") else c_ct
                ct = c_ct
        names, gem, ty, sizes = parse_name(raw, cat)
        sheet = code_sheet.get(code, "")
        coll = SHEET_COLLECTION.get(sheet, "dia")
        gemgrp = "pearl" if (gem and "パール" in gem[0]) else ("dia" if gem and "ダイヤ" in gem[0] else coll)
        products[code] = {
            "code": code, "raw": raw, "cat": cat, "group": CAT_GROUP.get(cat, "other"),
            "mat": mat, "ct": ct, "price": price, "stock": 1,
            "names": names, "spec": spec_line(gem, ty, mat, ct, sizes),
            "gem": gemgrp, "new": sheet == "写真",
            "high": bool(price and price >= 500000),
            "sizes": sizes,
            "gem_names": {"ja": gem[0], "en": gem[1], "zh": gem[2]} if gem else None,
            "badge": badge_of(sheet, names["ja"], gemgrp, price),
        }
    out = list(products.values())
    os.makedirs(os.path.join(ROOT, "data"), exist_ok=True)
    with open(os.path.join(ROOT, "data", "products.json"), "w") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    hi = sum(1 for p in out if p["high"]); nw = sum(1 for p in out if p["new"])
    pearls = sum(1 for p in out if p["gem"] == "pearl"); dias = sum(1 for p in out if p["gem"] == "dia")
    print(f"{len(out)} products | 新作 {nw} | ハイ(≥40万) {hi} | dia {dias} | pearl {pearls}")
    for p in out[:6] + out[60:63]:
        print(p["code"], "|", p["names"]["ja"], "|", p["spec"]["ja"], "|", p["price"])

if __name__ == "__main__":
    main()
